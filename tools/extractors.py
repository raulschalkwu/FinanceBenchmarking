#!/usr/bin/env python3
"""Rohdatei/URL -> Text. Gemeinsam genutzt von der GUI (tools/gui.py) und dem
Shepherd-MCP (tools/shepherd_mcp.py), damit „hochladen" überall gleich
funktioniert – ob per Weboberfläche oder per Chat mit einem Agenten.

Unterstützt: PDF (pypdf), DOCX (zipfile), XLSX/XLSM (openpyxl, Struktur statt
Rohzellen), HTML (stdlib-Parser), Text/Markdown; URLs per urllib.
"""
from __future__ import annotations
import html
import html.parser
import re
from pathlib import Path


class _HTMLText(html.parser.HTMLParser):
    """Minimaler HTML->Text-Extraktor (stdlib). Ignoriert script/style/nav/
    footer, setzt Absatzumbrüche bei Blockelementen."""
    _SKIP = {"script", "style", "noscript", "head", "nav", "footer", "aside"}
    _BREAK = {"p", "br", "div", "li", "tr", "h1", "h2", "h3", "h4", "section"}

    def __init__(self):
        super().__init__()
        self.parts: list[str] = []
        self._skip = 0

    def handle_starttag(self, tag, attrs):
        if tag in self._SKIP:
            self._skip += 1
        elif tag in self._BREAK:
            self.parts.append("\n")

    def handle_endtag(self, tag):
        if tag in self._SKIP and self._skip:
            self._skip -= 1
        elif tag in self._BREAK:
            self.parts.append("\n")

    def handle_data(self, data):
        if not self._skip and data.strip():
            self.parts.append(data)


def html_to_text(raw: bytes | str) -> str:
    s = raw.decode("utf-8", "ignore") if isinstance(raw, bytes) else raw
    p = _HTMLText()
    p.feed(s)
    return re.sub(r"\n{3,}", "\n\n", "".join(p.parts)).strip()


def extract_docx(data: bytes) -> str:
    """.docx = ZIP mit word/document.xml. Stdlib-only (zipfile)."""
    import io
    import zipfile
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        xml = z.read("word/document.xml").decode("utf-8", "ignore")
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"<[^>]+>", "", xml)
    return re.sub(r"\n{3,}", "\n\n", html.unescape(xml)).strip()


def extract_xlsx(data: bytes) -> str:
    """.xlsx -> kompakte STRUKTUR-Beschreibung (Blätter, Spalten, Zeilenzahl,
    wenige Beispielzeilen). Bewusst KEINE Rohzellen-Wüste."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        cols = [str(c) for c in header if c is not None] if header else []
        sample = []
        for i, r in enumerate(rows):
            if i >= 3:
                break
            sample.append(" | ".join("" if c is None else str(c) for c in r))
        out.append(
            f"Blatt \"{ws.title}\": {ws.max_row} Zeilen × {ws.max_column} Spalten\n"
            f"Spalten: {', '.join(cols) or '(keine Kopfzeile)'}\n"
            + ("Beispielzeilen:\n  " + "\n  ".join(sample) if sample else ""))
    wb.close()
    return "\n\n".join(out)


def extract_text(fname: str, data: bytes) -> tuple[str, str | None]:
    """Rohdatei -> Text. Gibt (text, warnung) zurück. Endung entscheidet."""
    low = fname.lower()
    if low.endswith((".xlsx", ".xlsm")):
        try:
            text = extract_xlsx(data)
            return (text, None) if text.strip() else ("", "Excel-Datei ist leer.")
        except ImportError:
            return "", "openpyxl fehlt: .venv/bin/pip install openpyxl"
        except Exception as e:
            return "", f"Excel-Extraktion fehlgeschlagen: {e}"
    if low.endswith(".pdf"):
        try:
            import io
            import pypdf
            reader = pypdf.PdfReader(io.BytesIO(data))
            pages = [(p.extract_text() or "").strip() for p in reader.pages]
            text = "\n\n".join(pg for pg in pages if pg)
            if not text.strip():
                return "", ("PDF enthält keinen extrahierbaren Text "
                            "(vermutlich gescannt/Bild – OCR nötig).")
            return text, None
        except ImportError:
            return "", "pypdf fehlt: .venv/bin/pip install pypdf"
        except Exception as e:
            return "", f"PDF-Extraktion fehlgeschlagen: {e}"
    if low.endswith(".docx"):
        try:
            text = extract_docx(data)
            return (text, None) if text.strip() else ("", "DOCX enthält keinen Text.")
        except Exception as e:
            return "", f"DOCX-Extraktion fehlgeschlagen: {e}"
    if low.endswith((".html", ".htm")):
        text = html_to_text(data)
        return (text, None) if text.strip() else ("", "HTML enthält keinen Text.")
    return data.decode("utf-8", "ignore"), None


def fetch_url(url: str) -> tuple[str, str, str | None]:
    """URL abrufen -> (text, quellname, warnung). HTML->Text, PDF per pypdf.
    Nur http/https. Der Nutzer/Agent gibt die URL selbst vor."""
    import urllib.request
    from urllib.parse import urlparse
    if not re.match(r"^https?://", url, re.I):
        return "", url, "Nur http(s)-URLs erlaubt."
    try:
        req = urllib.request.Request(
            url, headers={"User-Agent": "Mozilla/5.0 (VaultBot research fetch)"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            ctype = resp.headers.get("Content-Type", "").lower()
            data = resp.read(25 * 1024 * 1024)  # 25 MB Deckel
    except Exception as e:
        return "", url, f"Abruf fehlgeschlagen: {e}"
    name = Path(urlparse(url).path).name or urlparse(url).netloc or "web"
    if "pdf" in ctype or url.lower().endswith(".pdf"):
        text, warn = extract_text(name if name.endswith(".pdf") else name + ".pdf", data)
        return text, name, warn
    text = html_to_text(data)
    return (text, name, None) if text.strip() else ("", name, "Keine lesbaren Textinhalte.")


def extract_any(source: str) -> tuple[str, str, str | None]:
    """Universell: `source` ist eine URL, ein Dateipfad ODER direkt Rohtext.
    Gibt (text, quellname, warnung) zurück. Für den Chat-/Agenten-Weg."""
    s = source.strip()
    if re.match(r"^https?://", s, re.I):
        return fetch_url(s)
    p = Path(s)
    if len(s) < 400 and p.exists() and p.is_file():
        text, warn = extract_text(p.name, p.read_bytes())
        return text, p.name, warn
    # sonst: schon Rohtext
    return s, "text", None
